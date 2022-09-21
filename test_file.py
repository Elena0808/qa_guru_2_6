import os
import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_writing_to_zip():
    z = zipfile.ZipFile('./resources/test_zip.zip', 'w')
    for folder, subfolders, files in os.walk('./resources'):
        for file in files:
            if file.endswith('.pdf') or file.endswith('.csv') or file.endswith('.xlsx'):
                z.write(os.path.join(folder, file), file)
    z.close()


def test_read_csv():
    file = zipfile.ZipFile('./resources/test_zip.zip', 'r')
    csv_data = file.read('qa_test.csv').decode('utf-8')
    assert csv_data == 'Test'


def test_read_pdf():
    pdf_file = zipfile.ZipFile('./resources/test_zip.zip')
    zip_pdf = pdf_file.open('qa_test.pdf', 'r')
    pdf_read = PdfReader(zip_pdf)
    assert pdf_read.numPages == 1
    page = pdf_read.pages[0]
    text = page.extract_text()
    assert 'Python' in text


def test_read_xlsx():
    xlsx = zipfile.ZipFile('./resources/test_zip.zip').open('qa_test.xlsx')
    workbook = load_workbook(xlsx)
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == 'qa_3'
    assert sheet.cell(row=1, column=1).value == 'qa_test_1'
